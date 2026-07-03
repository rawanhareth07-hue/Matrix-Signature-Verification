"""
Report generator module for AI Signature Verification & Anti-Forgery System.

Generates comprehensive summary reports in PDF (using fpdf2), Excel (using openpyxl),
and CSV formats. Captures statistics, registered signatures, and verification history logs.
"""

import os
import csv
from datetime import datetime
from typing import Dict, List, Any
import pandas as pd
from fpdf import FPDF
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from config import RESULTS_DIR, logger


class ReportGenerator:
    """Creates database reports summarizing verification statistics and history.

    Supports exporting to PDF (with structured tables and system information),
    Excel (with multi-tab workbook formatting), and raw CSV logs.
    """

    def __init__(self, db_manager) -> None:
        """Initialize the ReportGenerator with a DatabaseManager instance.

        Args:
            db_manager: An initialized DatabaseManager object.
        """
        self.db = db_manager
        logger.debug("ReportGenerator initialized.")

    def generate_pdf_report(self, output_path: str = None) -> str:
        """Create a professional PDF report summarizing the system state and history.

        Args:
            output_path: Target PDF filepath. Defaults to config.RESULTS_DIR/report_<date>.pdf.

        Returns:
            The filepath to the generated PDF.
        """
        if output_path is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = os.path.join(RESULTS_DIR, f"verification_report_{timestamp}.pdf")

        try:
            # Fetch statistics and history
            stats = self.db.get_statistics()
            history = self.db.get_verification_history()

            # Create PDF instance
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font("Helvetica", size=12)

            # --- Title Banner ---
            pdf.set_fill_color(27, 110, 243)  # Primary blue (#1B6EF3)
            pdf.rect(0, 0, 210, 40, "F")
            pdf.set_text_color(255, 255, 255)
            pdf.set_font("Helvetica", "B", 18)
            pdf.text(15, 25, "AI Signature Verification System")
            pdf.set_font("Helvetica", size=10)
            pdf.text(15, 32, f"System Authentication Report - Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

            # Spacer
            pdf.ln(35)
            pdf.set_text_color(0, 0, 0)

            # --- Section 1: Dashboard Statistics ---
            pdf.set_font("Helvetica", "B", 14)
            pdf.set_text_color(27, 110, 243)
            pdf.cell(0, 10, "1. Executive Summary & KPIs", ln=True)
            pdf.set_text_color(0, 0, 0)
            pdf.set_font("Helvetica", size=10)
            pdf.line(10, pdf.get_y(), 200, pdf.get_y())
            pdf.ln(4)

            # Print KPI Cards (represented as structured table cells in PDF)
            col_w = 45
            row_h = 8
            # Header
            pdf.set_fill_color(240, 244, 250)
            pdf.set_font("Helvetica", "B", 9)
            pdf.cell(col_w, row_h, " Metric", border=1, fill=True)
            pdf.cell(col_w, row_h, " Value", border=1, fill=True)
            pdf.cell(col_w, row_h, " Metric", border=1, fill=True)
            pdf.cell(col_w, row_h, " Value", border=1, ln=True, fill=True)

            pdf.set_font("Helvetica", size=9)
            
            # Row 1
            pdf.cell(col_w, row_h, " Total Registered", border=1)
            pdf.cell(col_w, row_h, f" {stats['total_signatures']}", border=1)
            pdf.cell(col_w, row_h, " Total Verifications", border=1)
            pdf.cell(col_w, row_h, f" {stats['total_verifications']}", border=1, ln=True)

            # Row 2
            pdf.cell(col_w, row_h, " Genuine Count", border=1)
            pdf.cell(col_w, row_h, f" {stats['genuine_count']}", border=1)
            pdf.cell(col_w, row_h, " Forged Count", border=1)
            pdf.cell(col_w, row_h, f" {stats['forged_count']}", border=1, ln=True)

            # Row 3
            pdf.cell(col_w, row_h, " Accuracy Metric", border=1)
            pdf.cell(col_w, row_h, f" {stats['accuracy'] * 100:.1f}%", border=1)
            pdf.cell(col_w, row_h, " Avg. Similarity", border=1)
            pdf.cell(col_w, row_h, f" {stats['avg_similarity'] * 100:.1f}%", border=1, ln=True)

            # Row 4
            pdf.cell(col_w, row_h, " Avg. Processing Time", border=1)
            pdf.cell(col_w, row_h, f" {stats['avg_processing_time']:.3f}s", border=1)
            pdf.cell(col_w, row_h, " Status", border=1)
            pdf.cell(col_w, row_h, " Active", border=1, ln=True)

            pdf.ln(10)

            # --- Section 2: Verification Logs ---
            pdf.set_font("Helvetica", "B", 14)
            pdf.set_text_color(27, 110, 243)
            pdf.cell(0, 10, "2. Recent Verification Attempts", ln=True)
            pdf.set_text_color(0, 0, 0)
            pdf.set_font("Helvetica", size=10)
            pdf.line(10, pdf.get_y(), 200, pdf.get_y())
            pdf.ln(4)

            # Table Header
            fields = ["ID", "Username", "Similarity", "Decision", "Latency", "Timestamp"]
            widths = [10, 45, 25, 25, 25, 60]
            
            pdf.set_fill_color(27, 110, 243)
            pdf.set_text_color(255, 255, 255)
            pdf.set_font("Helvetica", "B", 9)
            
            for i, field in enumerate(fields):
                pdf.cell(widths[i], row_h, f" {field}", border=1, fill=True)
            pdf.ln()

            pdf.set_text_color(0, 0, 0)
            pdf.set_font("Helvetica", size=8)

            fill_toggle = False
            for entry in history[:30]:  # Limit to last 30 for PDF length
                pdf.set_fill_color(248, 250, 252) if fill_toggle else pdf.set_fill_color(255, 255, 255)
                
                pdf.cell(widths[0], row_h, f" {entry['id']}", border=1, fill=True)
                pdf.cell(widths[1], row_h, f" {entry.get('user_name', 'N/A')}", border=1, fill=True)
                pdf.cell(widths[2], row_h, f" {entry['similarity_score'] * 100:.1f}%", border=1, fill=True)
                
                # Color code decision
                if entry["decision"] == "Genuine":
                    pdf.set_text_color(0, 150, 0) # Green
                else:
                    pdf.set_text_color(200, 0, 0) # Red
                pdf.cell(widths[3], row_h, f" {entry['decision']}", border=1, fill=True)
                
                pdf.set_text_color(0, 0, 0)
                pdf.cell(widths[4], row_h, f" {entry['processing_time']:.3f}s", border=1, fill=True)
                pdf.cell(widths[5], row_h, f" {entry['created_at']}", border=1, ln=True, fill=True)
                
                fill_toggle = not fill_toggle

            # Footer signature/notes
            pdf.ln(15)
            pdf.set_font("Helvetica", "I", 8)
            pdf.cell(0, 10, "End of authentication report. Confidential information for enterprise use only.", align="C")

            pdf.output(output_path)
            logger.info("PDF report successfully written to %s", output_path)
            return output_path

        except Exception as e:
            logger.error("Failed to generate PDF report: %s", e)
            raise

    def generate_excel_report(self, output_path: str = None) -> str:
        """Create a formatted multi-sheet Excel workbook with statistics and history.

        Args:
            output_path: Target Excel filepath. Defaults to config.RESULTS_DIR/report_<date>.xlsx.

        Returns:
            The filepath to the generated Excel file.
        """
        if output_path is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = os.path.join(RESULTS_DIR, f"verification_report_{timestamp}.xlsx")

        try:
            # Fetch data from SQLite
            stats = self.db.get_statistics()
            history = self.db.get_verification_history()
            signatures = self.db.get_all_signatures()

            # Create standard dataframes
            df_stats = pd.DataFrame([stats])
            df_history = pd.DataFrame(history)
            df_sigs = pd.DataFrame(signatures)

            # Drop blob vectors to keep spreadsheet lightweight and legible
            if not df_sigs.empty and "feature_vector" in df_sigs.columns:
                df_sigs = df_sigs.drop(columns=["feature_vector"])

            # Use openpyxl to write sheets with formatting
            wb = openpyxl.Workbook()
            
            # --- Sheet 1: Statistics ---
            ws_stats = wb.active
            ws_stats.title = "System Summary"
            ws_stats.views.sheetView[0].showGridLines = True

            # Styling definitions
            font_title = Font(name="Calibri", size=16, bold=True, color="1B6EF3")
            font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            fill_header = PatternFill(start_color="1B6EF3", end_color="1B6EF3", fill_type="solid")
            thin_border = Border(
                left=Side(style='thin', color='DDDDDD'),
                right=Side(style='thin', color='DDDDDD'),
                top=Side(style='thin', color='DDDDDD'),
                bottom=Side(style='thin', color='DDDDDD')
            )

            # Write Title
            ws_stats["A1"] = "Executive Summary Statistics"
            ws_stats["A1"].font = font_title
            
            # Write key-value pairs
            row_idx = 3
            kpi_labels = {
                "total_signatures": "Total Registered Signatures",
                "total_verifications": "Total Verification Requests",
                "genuine_count": "Verified Genuine Signatures",
                "forged_count": "Detected Forged Signatures",
                "accuracy": "System Verification Accuracy",
                "avg_similarity": "Average Core Similarity",
                "avg_processing_time": "Average Latency (seconds)"
            }

            for key, label in kpi_labels.items():
                ws_stats.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
                val = stats.get(key, 0)
                
                # Format percentages and floats
                cell = ws_stats.cell(row=row_idx, column=2, value=val)
                if key in ["accuracy", "avg_similarity"]:
                    cell.number_format = "0.0%"
                elif key == "avg_processing_time":
                    cell.number_format = "0.000"
                
                ws_stats.cell(row=row_idx, column=1).border = thin_border
                cell.border = thin_border
                row_idx += 1

            # Autoscale column widths
            ws_stats.column_dimensions["A"].width = 30
            ws_stats.column_dimensions["B"].width = 15

            # --- Sheet 2: Verification Logs ---
            ws_logs = wb.create_sheet(title="Verification Logs")
            ws_logs.views.sheetView[0].showGridLines = True
            
            if not df_history.empty:
                # Rename/Clean columns
                cols_to_write = ["id", "signature_id", "similarity_score", "decision", "method", "processing_time", "created_at"]
                # Write Header
                for col_idx, col_name in enumerate(cols_to_write, 1):
                    cell = ws_logs.cell(row=1, column=col_idx, value=col_name.replace("_", " ").title())
                    cell.font = font_header
                    cell.fill = fill_header
                    cell.alignment = Alignment(horizontal="center")

                # Write Rows
                for row_idx, row in enumerate(df_history[cols_to_write].values, 2):
                    for col_idx, value in enumerate(row, 1):
                        cell = ws_logs.cell(row=row_idx, column=col_idx, value=value)
                        cell.border = thin_border
                        
                        # Formatting
                        if col_idx == 3:  # Similarity score
                            cell.number_format = "0.0%"
                        elif col_idx == 6:  # Latency
                            cell.number_format = "0.000"

                # Adjust column widths
                for col in ws_logs.columns:
                    max_len = max(len(str(cell.value or '')) for cell in col)
                    ws_logs.column_dimensions[col[0].column_letter].width = max(max_len + 3, 12)
            else:
                ws_logs.cell(row=1, column=1, value="No verification logs recorded.")

            # --- Sheet 3: Registered Signatures ---
            ws_sigs = wb.create_sheet(title="Registered Signatures")
            ws_sigs.views.sheetView[0].showGridLines = True
            
            if not df_sigs.empty:
                cols_to_write = ["id", "user_name", "image_path", "notes", "created_at"]
                # Write Header
                for col_idx, col_name in enumerate(cols_to_write, 1):
                    cell = ws_sigs.cell(row=1, column=col_idx, value=col_name.replace("_", " ").title())
                    cell.font = font_header
                    cell.fill = fill_header
                    cell.alignment = Alignment(horizontal="center")

                # Write Rows
                for row_idx, row in enumerate(df_sigs[cols_to_write].values, 2):
                    for col_idx, value in enumerate(row, 1):
                        cell = ws_sigs.cell(row=row_idx, column=col_idx, value=value)
                        cell.border = thin_border

                # Adjust column widths
                for col in ws_sigs.columns:
                    max_len = max(len(str(cell.value or '')) for cell in col)
                    ws_sigs.column_dimensions[col[0].column_letter].width = max(max_len + 3, 12)
            else:
                ws_sigs.cell(row=1, column=1, value="No signatures registered in database.")

            # Save file
            wb.save(output_path)
            logger.info("Excel report successfully written to %s", output_path)
            return output_path

        except Exception as e:
            logger.error("Failed to generate Excel report: %s", e)
            raise

    def generate_csv_report(self, output_path: str = None) -> str:
        """Create a CSV format export of the complete verification history log.

        Args:
            output_path: Target CSV filepath. Defaults to config.RESULTS_DIR/report_<date>.csv.

        Returns:
            The filepath to the generated CSV.
        """
        if output_path is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = os.path.join(RESULTS_DIR, f"verification_report_{timestamp}.csv")

        try:
            history = self.db.get_verification_history()

            if not history:
                # Write header-only if history is empty
                with open(output_path, "w", newline="", encoding="utf-8") as f:
                    writer = csv.writer(f)
                    writer.writerow(["ID", "Signature ID", "Username", "Similarity Score", "Decision", "Method", "Processing Time", "Timestamp"])
                return output_path

            keys = ["id", "signature_id", "user_name", "similarity_score", "decision", "method", "processing_time", "created_at"]
            
            with open(output_path, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                # Header row
                writer.writerow([key.replace("_", " ").title() for key in keys])
                
                # Data rows
                for row in history:
                    writer.writerow([row.get(key, "") for key in keys])

            logger.info("CSV report successfully written to %s", output_path)
            return output_path

        except Exception as e:
            logger.error("Failed to generate CSV report: %s", e)
            raise
